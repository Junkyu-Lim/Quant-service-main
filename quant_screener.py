# =========================================================
# quant_screener.py  (v14 - COMPLETE RESTORE)
# ---------------------------------------------------------
# [FIXED] Restored missing 'preprocess_indicators' and all exports.
# [RESTORED] 100% of original v8 logic for indicators.
# [INTEGRATED] 5-Strategy scoring & Market Leaders.
# =========================================================

import sys
import logging
import re
from pathlib import Path

import numpy as np
import pandas as pd

import config

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, desc="", **kwargs):
        items = list(iterable)
        total = len(items)
        for i, item in enumerate(items):
            if i % max(1, total // 10) == 0:
                print(f"  {desc}: {i}/{total} ({i*100//total}%)")
            yield item
        print(f"  {desc}: {total}/{total} (100%)")

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("SCREENER")

DATA_DIR = config.DATA_DIR

# ─────────────────────────────────────────────
# 계정 매핑 (v8 원본 완벽 복구)
# ─────────────────────────────────────────────
EXACT_ACCOUNTS = {
    "매출액": ["매출액", "영업수익", "이자수익", "보험료수익", "순영업수익"],
    "영업이익": ["영업이익"],
    "순이익": ["지배주주순이익", "당기순이익"],
    # 지배주주지분 최우선: 우선주·비지배지분 왜곡 방지
    "자본": ["지배주주지분", "지배기업주주지분", "자본", "자본총계"],
    "부채": ["부채", "부채총계"],
    "배당금": ["주당배당금"],
    "영업CF": [
        "영업활동현금흐름", "영업활동으로인한현금흐름",
        "영업활동 현금흐름", "영업활동으로 인한 현금흐름",
        "영업활동에서창출된현금", "영업에서창출된현금흐름",
    ],
    "투자CF": ["투자활동현금흐름", "투자활동으로인한현금흐름", "투자활동 현금흐름"],
    "CAPEX": [
        "유형자산의취득", "유형자산취득",
        "(-)유형자산의취득", "(-)유형자산취득",
        "유형자산 취득", "유형자산의 취득",
        "유형자산매입",
        "유형자산의증가", "(-)유형자산의증가",
    ],
    "자산총계": ["자산총계", "자산"],
    "유동자산": ["유동자산"],
    "유동부채": ["유동부채"],
    "매출총이익": ["매출총이익"],
    # 이자보상배율 산출용
    "이자비용": ["이자비용", "이자비용(금융비용)", "금융비용", "이자및할인비용"],
    # ROIC 초과현금 차감용
    "현금및현금성자산": ["현금및현금성자산", "현금및예치금", "현금및예금"],
    "단기금융상품": ["단기금융상품", "단기투자자산", "단기금융자산"],
}

EXCLUDE_KEYWORDS = ["증가율", "(-1Y)", "(평균)", "률(", "비율", "배율", "(-1A", "(-1Q", "/ 수정평균", "잉여금", "조정",
                    "연율화", "자본금"]
_UNIT_SUFFIX_RE = re.compile(r'\((?:억원|백만원|원|천억원|조원)\)$')

_ACCOUNT_MATCH_CACHE = {}
_YOY_MATCH_CACHE = {}

def _normalize_account(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'^\([+\-±]\)\s*', '', name)
    name = name.replace(' ', '').replace('\n', '').replace('\t', '')
    return name

def get_account_match(raw_name, target_key):
    cache_key = (raw_name, target_key)
    if cache_key in _ACCOUNT_MATCH_CACHE:
        return _ACCOUNT_MATCH_CACHE[cache_key]
    
    targets = EXACT_ACCOUNTS.get(target_key, [target_key])
    norm_targets = {_normalize_account(t) for t in targets}
    priority_map = {_normalize_account(t): i for i, t in enumerate(targets)}
    
    raw_str = str(raw_name)
    
    # Step 1: exact match
    if raw_str in targets:
        prio = priority_map.get(_normalize_account(raw_str), 999)
        res = (1, prio)
        _ACCOUNT_MATCH_CACHE[cache_key] = res
        return res
        
    norm_raw = _normalize_account(raw_str)
    
    # Step 2: normalized match
    if norm_raw in norm_targets:
        prio = priority_map.get(norm_raw, 999)
        res = (2, prio)
        _ACCOUNT_MATCH_CACHE[cache_key] = res
        return res
        
    # Step 2.5: 단위접미사 제거 후 재매칭
    stripped = _UNIT_SUFFIX_RE.sub('', norm_raw)
    if stripped in norm_targets:
        prio = priority_map.get(stripped, 999)
        res = (3, prio)
        _ACCOUNT_MATCH_CACHE[cache_key] = res
        return res
        
    # Step 3: startswith fallback
    if not any(kw in raw_str for kw in EXCLUDE_KEYWORDS):
        for nt in norm_targets:
            if norm_raw.startswith(nt):
                prio = priority_map.get(nt, 999)
                res = (4, prio)
                _ACCOUNT_MATCH_CACHE[cache_key] = res
                return res
                
    res = (0, 999)
    _ACCOUNT_MATCH_CACHE[cache_key] = res
    return res

def get_yoy_match(raw_name, target_key):
    cache_key = (raw_name, target_key)
    if cache_key in _YOY_MATCH_CACHE:
        return _YOY_MATCH_CACHE[cache_key]
        
    base_keys = EXACT_ACCOUNTS.get(target_key, [target_key])
    raw_str = str(raw_name)
    norm_name = _normalize_account(raw_str)
    
    is_match = False
    for bk in base_keys:
        norm_bk = _normalize_account(bk)
        if norm_name.startswith(norm_bk + "증가율"):
            is_match = True
            break
            
    _YOY_MATCH_CACHE[cache_key] = is_match
    return is_match

# ═════════════════════════════════════════════
# 유틸리티 (v8 원본 로직)
# ═════════════════════════════════════════════

def normalize_code(x):
    try:
        if pd.isna(x) or str(x).strip() == "": return np.nan
        s = str(x).strip()
        if '.' in s: s = s.split('.')[0]
        return s.zfill(6)
    except: return np.nan

def load_table(prefix: str) -> pd.DataFrame:
    import db as _db
    df = _db.load_latest(prefix)
    if df.empty: return df
    df.columns = df.columns.str.strip()
    if "종목코드" in df.columns:
        df["종목코드"] = df["종목코드"].apply(normalize_code)
        df = df.dropna(subset=["종목코드"])
    if "기준일" in df.columns: df["기준일"] = df["기준일"].astype(str).str[:10]
    for col in ["값", "종가", "시가총액", "상장주식수"]:
        if col in df.columns: df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

def find_account_value(df, target_key, date_filter=None):
    if df.empty or "계정" not in df.columns: return {}
    
    matched_rows = []
    
    if date_filter is not None:
        valid_dates = set(date_filter)
    else:
        valid_dates = None
        
    for row in df.itertuples(index=False):
        raw_name = row.계정
        match_level, priority = get_account_match(raw_name, target_key)
        
        if match_level > 0:
            date_val = str(row.기준일)
            if valid_dates and date_val not in valid_dates:
                continue
            val = row.값
            matched_rows.append((match_level, priority, date_val, val))
            
    if not matched_rows:
        return {}
        
    min_level = min(r[0] for r in matched_rows)
    best_rows = [r for r in matched_rows if r[0] == min_level]
    best_rows.sort(key=lambda x: (x[2], x[1]))
    
    result = {}
    for r in best_rows:
        date_str = r[2]
        if date_str not in result:
            val = r[3]
            result[date_str] = float(val) if pd.notna(val) else None
            
    return result

def preprocess_indicators(ind_df):
    if ind_df.empty: return ind_df
    return ind_df.drop_duplicates(["종목코드", "기준일", "계정", "지표구분"], keep="first")

def detect_unit_multiplier(ind_df):
    sam = ind_df[ind_df["종목코드"] == "005930"]
    if sam.empty: return 100_000_000
    rev_dict = find_account_value(sam[sam["지표구분"] == "RATIO_Y"], "매출액")
    if not rev_dict: return 100_000_000
    valid_revs = [v for v in rev_dict.values() if v is not None]
    if not valid_revs: return 100_000_000
    latest_rev = max(valid_revs)
    if latest_rev > 1e14: return 1
    elif latest_rev > 1e8: return 1_000_000
    else: return 100_000_000

# ═════════════════════════════════════════════
# 분석 유틸리티 (v8 원본)
# ═════════════════════════════════════════════

def calc_cagr(series_dict, min_years=2):
    if not series_dict or len(series_dict) < min_years: return np.nan
    dates = sorted(series_dict.keys())
    v0, v1 = series_dict[dates[0]], series_dict[dates[-1]]
    if v0 is None or v1 is None or v0 <= 0 or v1 <= 0: return np.nan
    try:
        years = (pd.Timestamp(dates[-1]) - pd.Timestamp(dates[0])).days / 365.25
        return ((v1 / v0) ** (1 / years) - 1) * 100 if years > 0.5 else np.nan
    except: return np.nan

def count_consecutive_growth(series_dict):
    if not series_dict or len(series_dict) < 2: return 0
    vals = [series_dict[d] for d in sorted(series_dict.keys())]
    # 최신 미보고 기간(None)은 연속성장 카운트에서 제외
    while vals and vals[-1] is None:
        vals.pop()
    if len(vals) < 2: return 0
    count = 0
    for i in range(len(vals) - 1, 0, -1):
        if vals[i] is None or vals[i - 1] is None: break
        if vals[i] > vals[i - 1]: count += 1
        else: break
    return count

def _read_yoy_from_ratio_q(q_data, key):
    """RATIO_Q 테이블의 YoY% 계정(증가율)을 직접 읽어 {날짜: yoy%} 반환.
    계정명 패턴: '{key}증가율(({key} / {key}(-1Y)) - 1) * 100 ...'"""
    if q_data.empty or "계정" not in q_data.columns:
        return {}
        
    matched_rows = []
    for row in q_data.itertuples(index=False):
        raw_name = row.계정
        if get_yoy_match(raw_name, key):
            matched_rows.append((str(row.기준일), row.값))
            
    if not matched_rows:
        return {}
        
    result = {}
    for r in matched_rows:
        date_str = r[0]
        val = r[1]
        if date_str not in result:
            if pd.notna(val):
                result[date_str] = float(val)
                
    return result


def calc_quarterly_yoy(q_data, key):
    res = {"latest_yoy": np.nan, "consecutive_yoy_growth": 0, "latest_quarter": "", "yoy_series": {}}
    vals = find_account_value(q_data, key)
    yoy_s = {}
    if len(vals) >= 5:
        for d in sorted(vals.keys()):
            prev_d = str(int(d[:4])-1) + d[4:]
            if prev_d in vals and vals[prev_d] is not None and vals[d] is not None and vals[prev_d] > 0:
                yoy_s[d] = ((vals[d]/vals[prev_d])-1)*100
    # 절대값으로 YoY 3개 미만이면 RATIO_Q 증가율 계정 직접 사용
    if len(yoy_s) < 3:
        yoy_direct = _read_yoy_from_ratio_q(q_data, key)
        if len(yoy_direct) >= len(yoy_s):
            yoy_s = yoy_direct
    if not yoy_s: return res
    res["yoy_series"], res["latest_quarter"] = yoy_s, max(yoy_s.keys())
    res["latest_yoy"] = yoy_s[res["latest_quarter"]]
    for d in sorted(yoy_s.keys(), reverse=True):
        if yoy_s[d] > 0: res["consecutive_yoy_growth"] += 1
        else: break
    return res

def calc_ttm_yoy(q_data, key):
    res = {"ttm_current": np.nan, "ttm_prev": np.nan, "ttm_yoy": np.nan}
    vals = find_account_value(q_data, key)
    if len(vals) < 4: return res
    dates = sorted(vals.keys())
    last4 = dates[-4:]
    prev4 = dates[-8:-4] if len(dates) >= 8 else []
    last4_valid = [d for d in last4 if d in vals and vals[d] is not None]
    prev4_valid = [d for d in prev4 if d in vals and vals[d] is not None]
    ttm_curr = sum(vals[d] for d in last4_valid)
    ttm_prev = sum(vals[d] for d in prev4_valid)
    if len(last4_valid) == 4: res["ttm_current"] = ttm_curr
    if len(prev4_valid) == 4: res["ttm_prev"] = ttm_prev
    if pd.notna(res["ttm_current"]) and pd.notna(res["ttm_prev"]) and res["ttm_prev"] > 0:
        res["ttm_yoy"] = ((res["ttm_current"] / res["ttm_prev"]) - 1) * 100
    return res

# ═════════════════════════════════════════════
# 메인 분석 엔진 (v8 원본 완벽 복구)
# ═════════════════════════════════════════════

def analyze_one_stock(ticker, ind_grp, fs_grp):
    res = {"종목코드": ticker}
    has_ind, has_fs = not ind_grp.empty, not fs_grp.empty
    
    if has_ind:
        q_data = ind_grp[ind_grp["지표구분"] == "RATIO_Q"]
        y_data = ind_grp[ind_grp["지표구분"] == "RATIO_Y"]
        hl_data = ind_grp[ind_grp["지표구분"].isin(["HIGHLIGHT", "HIGHLIGHT_E", "FORWARD_Y", "FORWARD_Q"])]

        _fs_y = fs_grp[fs_grp["주기"] == "y"] if has_fs and "주기" in fs_grp.columns else pd.DataFrame()

        def _yr(key):
            v = find_account_value(y_data, key)
            if not v:
                v = find_account_value(hl_data, key)
            if not v and not _fs_y.empty:
                v = find_account_value(_fs_y, key)
            return v

        _fs_q = fs_grp[fs_grp["주기"] == "q"] if has_fs and "주기" in fs_grp.columns else pd.DataFrame()

        for label, key in [("매출", "매출액"), ("영업이익", "영업이익"), ("순이익", "순이익")]:
            qyoy = calc_quarterly_yoy(q_data, key)
            if not qyoy["yoy_series"] and not _fs_q.empty:
                qyoy = calc_quarterly_yoy(_fs_q, key)
            res[f"Q_{label}_YoY(%)"] = qyoy["latest_yoy"]
            res[f"Q_{label}_연속YoY성장"] = qyoy["consecutive_yoy_growth"]
            ttmy = calc_ttm_yoy(q_data, key)
            if pd.isna(ttmy["ttm_current"]) and not _fs_q.empty:
                ttmy = calc_ttm_yoy(_fs_q, key)
            res[f"TTM_{label}_YoY(%)"] = ttmy["ttm_yoy"]
            val = ttmy["ttm_current"]
            if pd.isna(val):
                y_vals = _yr(key)
                if y_vals: val = y_vals[max(y_vals.keys())]
            res[f"TTM_{label}"] = val
        res["최근분기"] = sorted(q_data["기준일"].unique())[-1] if not q_data.empty else ""

        # ── 실적 가속도 (Earnings Acceleration) ──
        # 영업이익/매출 각각 YoY의 미분값(ΔYoY) 계산, 2분기 연속 가속 여부 판별
        def _calc_acceleration(qyoy_result):
            """YoY 시리즈에서 최근 2개 ΔYoY 계산 → 2분기 연속 양(+) 여부 반환"""
            yoy_s = qyoy_result.get("yoy_series", {})
            if len(yoy_s) < 3:
                return np.nan, False
            dates = sorted(yoy_s.keys())
            # 최근 3분기 YoY로 최근 2개 ΔYoY 산출
            d0, d1, d2 = dates[-3], dates[-2], dates[-1]
            delta_prev = yoy_s[d1] - yoy_s[d0]
            delta_latest = yoy_s[d2] - yoy_s[d1]
            consecutive = (delta_prev > 0 and delta_latest > 0)
            return delta_latest, consecutive

        op_qyoy = calc_quarterly_yoy(q_data, "영업이익")
        if not op_qyoy["yoy_series"] and not _fs_q.empty:
            op_qyoy = calc_quarterly_yoy(_fs_q, "영업이익")
        rev_qyoy = calc_quarterly_yoy(q_data, "매출액")
        if not rev_qyoy["yoy_series"] and not _fs_q.empty:
            rev_qyoy = calc_quarterly_yoy(_fs_q, "매출액")
        op_accel, op_consec = _calc_acceleration(op_qyoy)
        rev_accel, rev_consec = _calc_acceleration(rev_qyoy)
        res["영업이익_가속도"] = op_accel
        res["매출_가속도"] = rev_accel

        # 4분기 빅배스 Fallback: 최신 분기가 Q4이고 영업이익 가속 꺾였어도
        # 매출 가속 유지 + GPM 훼손 없으면 실적가속_연속 = 1 인정 (GPM은 아래서 산출)
        # → 일단 연속 여부 잠정 저장, GPM 계산 후 최종 결정
        latest_q = res.get("최근분기", "")
        is_q4 = latest_q.endswith("12") or latest_q.endswith("Q4") or (
            len(latest_q) >= 6 and latest_q[4:6] == "12"
        )
        res["_op_consec_accel"] = op_consec
        res["_rev_consec_accel"] = rev_consec
        res["_is_q4"] = is_q4

        total_assets_s = _yr("자산총계")
        equity_s = _yr("자본")
        debt_s = _yr("부채")
        current_assets_s = _yr("유동자산")
        current_liab_s = _yr("유동부채")
        gross_profit_s = _yr("매출총이익")
        cash_s = _yr("현금및현금성자산")
        stfi_s = _yr("단기금융상품")
        interest_s = _yr("이자비용")
        res["자산총계"] = total_assets_s[max(total_assets_s.keys())] if total_assets_s else np.nan
        res["자본"] = equity_s[max(equity_s.keys())] if equity_s else np.nan
        res["부채"] = debt_s[max(debt_s.keys())] if debt_s else np.nan

        rev_s = _yr("매출액")
        op_s = _yr("영업이익")
        ni_s = _yr("순이익")
        res.update({
            "매출_CAGR": calc_cagr(rev_s), "영업이익_CAGR": calc_cagr(op_s), "순이익_CAGR": calc_cagr(ni_s),
            "매출_연속성장": count_consecutive_growth(rev_s), "영업이익_연속성장": count_consecutive_growth(op_s), "순이익_연속성장": count_consecutive_growth(ni_s),
        })
        res["데이터_연수"] = len(rev_s)

        if len(rev_s) >= 2 and len(op_s) >= 2:
            l, p = sorted(rev_s.keys())[-1], sorted(rev_s.keys())[-2]
            opm_l = (op_s[l]/rev_s[l]*100) if (rev_s[l] is not None and op_s.get(l) is not None and rev_s[l] > 0) else np.nan
            opm_p = (op_s[p]/rev_s[p]*100) if (rev_s[p] is not None and op_s.get(p) is not None and rev_s[p] > 0) else np.nan
            res["영업이익률_최근"], res["영업이익률_전년"] = opm_l, opm_p
            res["이익률_개선"] = 1 if pd.notna(opm_l) and pd.notna(opm_p) and opm_l > opm_p else 0
            delta = opm_l - opm_p if pd.notna(opm_l) and pd.notna(opm_p) else np.nan
            res["이익률_변동폭"], res["이익률_급개선"] = delta, (1 if (delta or 0) >= 5 else 0)
        else:
            res["영업이익률_최근"] = np.nan
            res["영업이익률_전년"] = np.nan
            res["이익률_개선"] = 0
            res["이익률_변동폭"] = np.nan
            res["이익률_급개선"] = 0

        if len(ni_s) >= 2:
            ni_vals = [ni_s[d] for d in sorted(ni_s.keys())]
            res["흑자전환"] = 1 if (ni_vals[-2] is not None and ni_vals[-1] is not None and ni_vals[-2] < 0 and ni_vals[-1] > 0) else 0
            res["순이익_당기양수"] = 1 if (ni_vals[-1] is not None and ni_vals[-1] > 0) else 0
            res["순이익_전년음수"] = 1 if (ni_vals[-2] is not None and ni_vals[-2] < 0) else 0
        else:
            res["흑자전환"] = 0
            res["순이익_당기양수"] = 0
            res["순이익_전년음수"] = 0

        ocf_s = find_account_value(y_data, "영업CF")
        if not ocf_s: ocf_s = find_account_value(ind_grp[ind_grp["지표구분"].isin(["HIGHLIGHT", "HIGHLIGHT_E", "FORWARD_Y", "FORWARD_Q"])], "영업CF")
        if not ocf_s and has_fs: ocf_s = find_account_value(fs_grp[fs_grp["주기"]=="y"], "영업CF")
        capex_s = find_account_value(y_data, "CAPEX")
        if not capex_s and has_fs: capex_s = find_account_value(fs_grp[fs_grp["주기"]=="y"], "CAPEX")
        capex_s = {d: abs(v) for d, v in capex_s.items() if v is not None}
        
        fcf_s = {d: (ocf_s[d] - capex_s[d]) for d in (set(ocf_s.keys()) & set(capex_s.keys())) if ocf_s[d] is not None and capex_s[d] is not None}
        ttm_ocf = ocf_s[max(ocf_s.keys())] if ocf_s else np.nan
        ttm_capex = capex_s[max(capex_s.keys())] if capex_s else np.nan
        res.update({
            "TTM_영업CF": ttm_ocf, "TTM_CAPEX": ttm_capex, "TTM_FCF": (ttm_ocf - ttm_capex if pd.notna(ttm_ocf) and pd.notna(ttm_capex) else np.nan),
            "영업CF_CAGR": calc_cagr(ocf_s), "FCF_CAGR": calc_cagr(fcf_s), "영업CF_연속성장": count_consecutive_growth(ocf_s)
        })

        f1 = 1 if (res.get("TTM_순이익") or 0) > 0 else 0
        f2 = 1 if (ttm_ocf or 0) > 0 else 0
        f4 = 1 if (ttm_ocf or 0) > (res.get("TTM_순이익") or 0) and f1 else 0

        # F3: ROA 개선 (ni/total_assets 최근 > 전년)
        f3 = 0
        if len(ni_s) >= 2 and len(total_assets_s) >= 2:
            ni_dates = sorted(ni_s.keys())
            ta_dates = sorted(total_assets_s.keys())
            if ni_dates[-1] in total_assets_s and ni_dates[-2] in total_assets_s:
                roa_cur = ni_s[ni_dates[-1]] / total_assets_s[ni_dates[-1]] if (total_assets_s[ni_dates[-1]] and ni_s[ni_dates[-1]] is not None) else 0
                roa_prev = ni_s[ni_dates[-2]] / total_assets_s[ni_dates[-2]] if (total_assets_s[ni_dates[-2]] and ni_s[ni_dates[-2]] is not None) else 0
                f3 = 1 if roa_cur > roa_prev else 0
            elif len(ta_dates) >= 2:
                roa_cur = ni_s[ni_dates[-1]] / total_assets_s[ta_dates[-1]] if (total_assets_s[ta_dates[-1]] and ni_s[ni_dates[-1]] is not None) else 0
                roa_prev = ni_s[ni_dates[-2]] / total_assets_s[ta_dates[-2]] if (total_assets_s[ta_dates[-2]] and ni_s[ni_dates[-2]] is not None) else 0
                f3 = 1 if roa_cur > roa_prev else 0

        # F5: 레버리지 감소 (부채비율 최근 < 전년)
        f5 = 0
        if len(debt_s) >= 2 and len(equity_s) >= 2:
            d_dates = sorted(debt_s.keys())
            e_dates = sorted(equity_s.keys())
            if len(d_dates) >= 2 and len(e_dates) >= 2:
                dr_cur = debt_s[d_dates[-1]] / equity_s[e_dates[-1]] if (equity_s[e_dates[-1]] and debt_s[d_dates[-1]] is not None) else 999
                dr_prev = debt_s[d_dates[-2]] / equity_s[e_dates[-2]] if (equity_s[e_dates[-2]] and debt_s[d_dates[-2]] is not None) else 999
                f5 = 1 if dr_cur < dr_prev else 0

        # F6: 유동성 개선 (유동비율 최근 > 전년)
        f6 = 0
        _cr_cur_ratio = np.nan  # 유동비율(%) 저장용
        if len(current_assets_s) >= 2 and len(current_liab_s) >= 2:
            ca_dates = sorted(current_assets_s.keys())
            cl_dates = sorted(current_liab_s.keys())
            if len(ca_dates) >= 2 and len(cl_dates) >= 2:
                cr_cur = current_assets_s[ca_dates[-1]] / current_liab_s[cl_dates[-1]] if (current_liab_s[cl_dates[-1]] and current_assets_s[ca_dates[-1]] is not None) else 0
                cr_prev = current_assets_s[ca_dates[-2]] / current_liab_s[cl_dates[-2]] if (current_liab_s[cl_dates[-2]] and current_assets_s[ca_dates[-2]] is not None) else 0
                f6 = 1 if cr_cur > cr_prev else 0
                if cr_cur > 0:
                    _cr_cur_ratio = cr_cur * 100  # 유동비율(%)

        # F7: 희석 없음 — shares 데이터 없을 시 자본/순이익 proxy로 판별
        # 기본값 1 (희석 없음 가정): 대부분 기업은 희석 안 함, 데이터 부재 시 패널티 금지
        f7 = 1
        if len(equity_s) >= 2 and len(ni_s) >= 2:
            eq_dates = sorted(equity_s.keys())
            eq_prev = equity_s[eq_dates[-2]]
            eq_cur = equity_s[eq_dates[-1]]
            ni_cur = ni_s.get(sorted(ni_s.keys())[-1])
            # 자본증가분이 순이익의 1.5배를 크게 초과하면 유증 의심 → F7=0
            if (eq_prev is not None and eq_cur is not None and
                    ni_cur is not None and eq_prev > 0):
                eq_growth = eq_cur - eq_prev
                if eq_growth > abs(ni_cur) * 1.5 + abs(eq_prev) * 0.05:
                    f7 = 0  # proxy: 유상증자 의심
        # calc_valuation에서 shares_df 실제 데이터로 재정의 가능

        # F8: 매출총이익률 개선
        f8 = 0
        if len(gross_profit_s) >= 2 and len(rev_s) >= 2:
            gp_dates = sorted(gross_profit_s.keys())
            rv_dates = sorted(rev_s.keys())
            if len(gp_dates) >= 2 and len(rv_dates) >= 2:
                gpm_cur = gross_profit_s[gp_dates[-1]] / rev_s[rv_dates[-1]] if (rev_s[rv_dates[-1]] and gross_profit_s[gp_dates[-1]] is not None) else 0
                gpm_prev = gross_profit_s[gp_dates[-2]] / rev_s[rv_dates[-2]] if (rev_s[rv_dates[-2]] and gross_profit_s[gp_dates[-2]] is not None) else 0
                f8 = 1 if gpm_cur > gpm_prev else 0

        # F9: 자산회전율 개선 (매출/자산총계 최근 > 전년)
        f9 = 0
        if len(rev_s) >= 2 and len(total_assets_s) >= 2:
            rv_dates = sorted(rev_s.keys())
            ta_dates = sorted(total_assets_s.keys())
            if len(rv_dates) >= 2 and len(ta_dates) >= 2:
                at_cur = rev_s[rv_dates[-1]] / total_assets_s[ta_dates[-1]] if (total_assets_s[ta_dates[-1]] and rev_s[rv_dates[-1]] is not None) else 0
                at_prev = rev_s[rv_dates[-2]] / total_assets_s[ta_dates[-2]] if (total_assets_s[ta_dates[-2]] and rev_s[rv_dates[-2]] is not None) else 0
                f9 = 1 if at_cur > at_prev else 0

        res["F스코어"] = f1 + f2 + f3 + f4 + f5 + f6 + f7 + f8 + f9
        res["F1_수익성"], res["F2_영업CF"], res["F3_ROA개선"], res["F4_이익품질"] = f1, f2, f3, f4
        res["F5_레버리지"], res["F6_유동성"], res["F7_희석없음"] = f5, f6, f7
        res["F8_매출총이익률"], res["F9_자산회전율"] = f8, f9
        res["유동비율(%)"] = _cr_cur_ratio

        # ── GPM 연속값 & ROIC ──
        gpm_cur_val = np.nan
        gpm_prev_val = np.nan
        if len(gross_profit_s) >= 2 and len(rev_s) >= 2:
            gp_dates = sorted(gross_profit_s.keys())
            rv_dates_y = sorted(rev_s.keys())
            if len(gp_dates) >= 2 and len(rv_dates_y) >= 2:
                gpm_cur_val = (gross_profit_s[gp_dates[-1]] / rev_s[rv_dates_y[-1]] * 100
                               if rev_s[rv_dates_y[-1]] and gross_profit_s[gp_dates[-1]] is not None else np.nan)
                gpm_prev_val = (gross_profit_s[gp_dates[-2]] / rev_s[rv_dates_y[-2]] * 100
                                if rev_s[rv_dates_y[-2]] and gross_profit_s[gp_dates[-2]] is not None else np.nan)
        gpm_delta = (gpm_cur_val - gpm_prev_val) if pd.notna(gpm_cur_val) and pd.notna(gpm_prev_val) else np.nan
        res["GPM_최근(%)"] = gpm_cur_val
        res["GPM_전년(%)"] = gpm_prev_val
        res["GPM_변화(pp)"] = gpm_delta

        # 4분기 빅배스 Fallback 최종 결정 (GPM 훼손 여부 반영)
        op_consec = res.pop("_op_consec_accel", False)
        rev_consec = res.pop("_rev_consec_accel", False)
        is_q4 = res.pop("_is_q4", False)
        gpm_ok = pd.notna(gpm_delta) and gpm_delta >= 0  # GPM 훼손 없음
        # OPM 교차검증: 연간 데이터 있을 때만 체크, 없으면 pass (분기가속 우선 신뢰)
        has_opm_data = pd.notna(res.get("영업이익률_최근")) and pd.notna(res.get("영업이익률_전년"))
        opm_ok = (not has_opm_data) or (res.get("이익률_개선", 0) == 1)
        if op_consec and opm_ok:
            res["실적가속_연속"] = 1
        elif is_q4 and rev_consec and gpm_ok:
            res["실적가속_연속"] = 1  # 4Q fallback: 매출 가속 + GPM 유지
        else:
            res["실적가속_연속"] = 0

        # ROIC 계산 (NOPAT / Invested Capital, 초과현금 차감 적용)
        # IC = 자산총계 - 유동부채 - 초과현금 (초과현금 = 현금성자산 + 단기금융상품 - 매출의 2.5%)
        roic_cur = np.nan
        roic_prev = np.nan
        if op_s and total_assets_s:
            op_dates_y = sorted(op_s.keys())
            ta_dates_y = sorted(total_assets_s.keys())
            cl_dates = sorted(current_liab_s.keys()) if current_liab_s else []
            rv_dates_y2 = sorted(rev_s.keys()) if rev_s else []
            cash_dates = sorted(cash_s.keys()) if cash_s else []
            stfi_dates = sorted(stfi_s.keys()) if stfi_s else []
            for i, (op_key, ta_key) in enumerate([
                (op_dates_y[-1], ta_dates_y[-1]) if len(op_dates_y) >= 1 and len(ta_dates_y) >= 1 else (None, None),
                (op_dates_y[-2], ta_dates_y[-2]) if len(op_dates_y) >= 2 and len(ta_dates_y) >= 2 else (None, None),
            ]):
                if op_key is None or ta_key is None:
                    continue
                op_val = op_s.get(op_key)
                ta_val = total_assets_s.get(ta_key)
                if op_val is None or ta_val is None or ta_val == 0:
                    continue
                cl_val = current_liab_s.get(cl_dates[-1 - i]) if cl_dates and len(cl_dates) > i else None
                ic = (ta_val - cl_val) if cl_val is not None else ta_val
                # 초과현금 차감: 현금성자산 + 단기금융상품 - 매출의 2.5% (운영필요현금)
                cash_val = cash_s.get(cash_dates[-1 - i]) if cash_dates and len(cash_dates) > i else None
                stfi_val = stfi_s.get(stfi_dates[-1 - i]) if stfi_dates and len(stfi_dates) > i else None
                rev_val = rev_s.get(rv_dates_y2[-1 - i]) if rv_dates_y2 and len(rv_dates_y2) > i else None
                if cash_val is not None or stfi_val is not None:
                    total_cash = (cash_val or 0) + (stfi_val or 0)
                    op_cash_need = (rev_val * 0.025) if rev_val is not None and rev_val > 0 else 0
                    excess_cash = max(0, total_cash - op_cash_need)
                    ic = ic - excess_cash
                if ic <= 0:
                    ic = ta_val  # floor: 초과차감으로 음수 방지
                nopat = op_val * (1 - 0.22)
                roic_val = nopat / ic * 100
                if i == 0:
                    roic_cur = roic_val
                else:
                    roic_prev = roic_val
        res["ROIC(%)"] = roic_cur
        res["ROIC_전년(%)"] = roic_prev
        res["ROIC_개선"] = 1 if pd.notna(roic_cur) and pd.notna(roic_prev) and roic_cur > roic_prev else 0

        # 퀄리티 턴어라운드 복합 신호
        # GPM +2%p 이상 AND 영업CF > 0 AND ROIC 개선
        ttm_ocf_val = res.get("TTM_영업CF")
        ocf_positive = pd.notna(ttm_ocf_val) and ttm_ocf_val > 0
        gpm_surge = pd.notna(gpm_delta) and gpm_delta >= 2.0
        res["퀄리티_턴어라운드"] = 1 if (gpm_surge and ocf_positive and res["ROIC_개선"] == 1) else 0

        # ── 이자보상배율 (TTM 영업이익 / 이자비용) ──
        ttm_interest = interest_s[max(interest_s.keys())] if interest_s else np.nan
        ttm_op_for_icr = res.get("TTM_영업이익")
        if pd.isna(ttm_op_for_icr):
            ttm_op_for_icr = op_s[max(op_s.keys())] if op_s else np.nan
        if pd.notna(ttm_op_for_icr) and pd.notna(ttm_interest) and ttm_interest != 0:
            res["이자보상배율"] = ttm_op_for_icr / abs(ttm_interest)
        else:
            # 이자비용 0 = 무차입, 이자보상 필요 없음 → 99로 표시
            res["이자보상배율"] = 99.0 if (pd.notna(ttm_interest) and ttm_interest == 0) else np.nan

        dps_s = find_account_value(ind_grp[ind_grp["지표구분"]=="DPS"], "배당금")
        # 미래 날짜 및 분기 기준일 제거: 연말(12,8,11월 등 결산월) + 오늘 이전만 사용
        today_str = pd.Timestamp.today().strftime("%Y-%m-%d")
        valid_dps_keys = [k for k in dps_s if str(k) <= today_str and pd.Timestamp(k).month not in (3, 6, 9)]
        annual_dps_s = {k: dps_s[k] for k in valid_dps_keys}
        res["DPS_최근"] = annual_dps_s[max(annual_dps_s.keys())] if annual_dps_s else np.nan
        res["DPS_CAGR"], res["배당_연속증가"] = calc_cagr(annual_dps_s), count_consecutive_growth(annual_dps_s)
        res["배당_수익동반증가"] = 1 if res["순이익_연속성장"] >= 2 and res["배당_연속증가"] >= 1 else 0

        # ── Forward 컨센서스 추정치 ──────────────────────────────────────────
        fwd_y_data = ind_grp[ind_grp["지표구분"] == "FORWARD_Y"]
        # 동적 Forward 연도 선택: RATIO_Y 최신 기준일 이후 연도만
        latest_ratio_date = ind_grp[ind_grp["지표구분"] == "RATIO_Y"]["기준일"].max() if not ind_grp[ind_grp["지표구분"] == "RATIO_Y"].empty else ""
        if not fwd_y_data.empty and latest_ratio_date:
            fwd_dates = sorted([d for d in fwd_y_data["기준일"].unique() if str(d) > str(latest_ratio_date)])
        elif not fwd_y_data.empty:
            fwd_dates = sorted(fwd_y_data["기준일"].unique())
        else:
            fwd_dates = []
        fwd_1yr = fwd_dates[0] if fwd_dates else None
        fwd_2yr = fwd_dates[1] if len(fwd_dates) >= 2 else None

        def _fwd_val(date, account_keyword):
            """FORWARD_Y 데이터에서 계정명 키워드로 값 추출"""
            if date is None or fwd_y_data.empty:
                return np.nan
            rows = fwd_y_data[fwd_y_data["기준일"] == date]
            if rows.empty:
                return np.nan
            # 계정명에 keyword가 포함된 행 찾기
            matched = rows[rows["계정"].str.contains(account_keyword, na=False, regex=False)]
            if matched.empty:
                return np.nan
            return matched.iloc[0]["값"]

        if fwd_1yr is not None:
            res["Fwd_PER"] = _fwd_val(fwd_1yr, "PER")
            res["Fwd_PBR"] = _fwd_val(fwd_1yr, "PBR")
            res["Fwd_EPS"] = _fwd_val(fwd_1yr, "EPS")
            res["Fwd_ROE(%)"] = _fwd_val(fwd_1yr, "ROE")
            res["Fwd_OPM(%)"] = _fwd_val(fwd_1yr, "영업이익률")
            # 실제 유효한 값이 하나라도 있어야 커버리지=1 (행만 존재하고 모두 NaN인 종목 제외)
            res["컨센서스_커버리지"] = 1 if pd.notna(res["Fwd_PER"]) or pd.notna(res["Fwd_ROE(%)"]) or pd.notna(res["Fwd_EPS"]) else 0
            # 성장률 계산 (TTM 대비)
            fwd_op = _fwd_val(fwd_1yr, "영업이익")
            fwd_rev = _fwd_val(fwd_1yr, "매출액")
            fwd_ni = _fwd_val(fwd_1yr, "지배주주순이익")
            if pd.isna(fwd_ni):
                fwd_ni = _fwd_val(fwd_1yr, "당기순이익")
            ttm_op = res.get("TTM_영업이익", np.nan)
            ttm_rev = res.get("TTM_매출", np.nan)
            ttm_ni = res.get("TTM_순이익", np.nan)
            res["Fwd_영업이익_성장률(%)"] = (fwd_op / ttm_op - 1) * 100 if pd.notna(fwd_op) and pd.notna(ttm_op) and ttm_op != 0 else np.nan
            res["Fwd_매출_성장률(%)"] = (fwd_rev / ttm_rev - 1) * 100 if pd.notna(fwd_rev) and pd.notna(ttm_rev) and ttm_rev != 0 else np.nan
            res["Fwd_순이익_성장률(%)"] = (fwd_ni / ttm_ni - 1) * 100 if pd.notna(fwd_ni) and pd.notna(ttm_ni) and ttm_ni != 0 else np.nan
            # 2년 Forward 성장
            if fwd_2yr is not None:
                fwd_op_2yr = _fwd_val(fwd_2yr, "영업이익")
                res["Fwd_2yr_영업이익_성장(%)"] = (fwd_op_2yr / fwd_op - 1) * 100 if pd.notna(fwd_op_2yr) and pd.notna(fwd_op) and fwd_op != 0 else np.nan
            else:
                res["Fwd_2yr_영업이익_성장(%)"] = np.nan
        else:
            res["컨센서스_커버리지"] = 0
            for col in ["Fwd_PER", "Fwd_PBR", "Fwd_EPS", "Fwd_ROE(%)", "Fwd_OPM(%)",
                        "Fwd_영업이익_성장률(%)", "Fwd_매출_성장률(%)", "Fwd_순이익_성장률(%)", "Fwd_2yr_영업이익_성장(%)"]:
                res[col] = np.nan

    return res

def analyze_all(fs_df, ind_df, progress_callback=None):
    """Analyze all stocks with optional progress callback.

    Args:
        fs_df: Financial statements dataframe
        ind_df: Indicators dataframe
        progress_callback: Optional callable(stage: str, pct: int) for progress tracking
    """
    results = []
    tickers = list(set(fs_df["종목코드"].unique()) | set(ind_df["종목코드"].unique()))
    total = len(tickers)
    
    log.info("사전 그룹화 진행 중 (Pre-grouping data for performance)...")
    ind_grouped = {k: v for k, v in ind_df.groupby("종목코드")} if not ind_df.empty else {}
    fs_grouped = {k: v for k, v in fs_df.groupby("종목코드")} if not fs_df.empty else {}
    empty_ind = pd.DataFrame(columns=ind_df.columns) if not ind_df.empty else pd.DataFrame()
    empty_fs = pd.DataFrame(columns=fs_df.columns) if not fs_df.empty else pd.DataFrame()

    for idx, ticker in enumerate(tqdm(tickers, desc="펀더멘털 분석")):
        ind_grp = ind_grouped.get(ticker, empty_ind)
        fs_grp = fs_grouped.get(ticker, empty_fs)
        results.append(analyze_one_stock(ticker, ind_grp=ind_grp, fs_grp=fs_grp))
        if progress_callback and idx % max(1, total // 10) == 0:
            pct = 62 + int((idx / total) * 8)  # 62% ~ 70% 범위
            progress_callback(f"펀더멘털 분석 중 ({idx}/{total})", pct)

    if progress_callback:
        progress_callback(f"펀더멘털 분석 완료 ({total}/{total})", 70)

    return pd.DataFrame(results)

# ═════════════════════════════════════════════
# 밸류에이션 & 기술적 지표 (v8 원본 + 수급)
# ═════════════════════════════════════════════

def calc_valuation(daily, anal_df, multiplier, shares_df):
    df = daily[["종목코드", "종목명", "종가", "시가총액", "상장주식수"]].drop_duplicates("종목코드").merge(anal_df, on="종목코드", how="inner")
    if shares_df is not None and not shares_df.empty:
        s_map = shares_df.drop_duplicates("종목코드").set_index("종목코드")["발행주식수"]
        mask = df["상장주식수"].isna() | (df["상장주식수"] == 0)
        df.loc[mask, "상장주식수"] = df.loc[mask, "종목코드"].map(s_map)

    M = multiplier
    df["PER"] = np.where((df["TTM_순이익"] > 0) & (df["시가총액"] > 0), df["시가총액"] / (df["TTM_순이익"] * M), np.nan)
    df["PBR"] = np.where((df["자본"] > 0) & (df["시가총액"] > 0), df["시가총액"] / (df["자본"] * M), np.nan)
    df["PSR"] = np.where((df["TTM_매출"] > 0) & (df["시가총액"] > 0), df["시가총액"] / (df["TTM_매출"] * M), np.nan)
    df["ROE(%)"] = np.where((df["자본"] > 0), (df["TTM_순이익"] / df["자본"]) * 100, np.nan)
    df["부채비율(%)"] = np.where((df["자본"] > 0), (df["부채"] / df["자본"]) * 100, np.nan)
    df["영업이익률(%)"] = df["영업이익률_최근"]
    df["배당수익률(%)"] = np.where((df["종가"] > 0) & (df["DPS_최근"] > 0), (df["DPS_최근"] / df["종가"]) * 100, 0)
    # PEG: 일회성 이익 급증 방지를 위해 CAGR 상한 100% 적용
    cagr_capped = np.minimum(df["순이익_CAGR"], 100)
    df["PEG"] = np.where((df["PER"] > 0) & (cagr_capped > 0), df["PER"] / cagr_capped, np.nan)
    df["FCF수익률(%)"] = np.where((df["시가총액"] > 0) & (df["TTM_FCF"] != 0), (df["TTM_FCF"] * M / df["시가총액"]) * 100, np.nan)
    df["이익수익률(%)"] = np.where((df["시가총액"] > 0) & (df["TTM_순이익"] > 0), (df["TTM_순이익"] * M / df["시가총액"]) * 100, np.nan)
    df["현금전환율(%)"] = np.where(pd.notna(df["TTM_영업CF"]) & (df["TTM_순이익"] > 0), (df["TTM_영업CF"] / df["TTM_순이익"]) * 100, np.nan)
    df["CAPEX비율(%)"] = np.where(pd.notna(df["TTM_CAPEX"]) & (df["TTM_영업CF"] > 0), (df["TTM_CAPEX"] / df["TTM_영업CF"]) * 100, np.nan)
    df["부채상환능력"] = np.where((df["TTM_영업CF"] > 0) & (df["부채"] > 0), df["TTM_영업CF"] / df["부채"], np.nan)
    df["이익품질_양호"] = np.where((df["TTM_영업CF"] > df["TTM_순이익"]) & (df["TTM_순이익"] > 0), 1, 0)

    shares = df["상장주식수"].replace(0, np.nan)
    df["BPS"], df["EPS"] = (df["자본"] * M) / shares, (df["TTM_순이익"] * M) / shares

    # 배당성향(%) — EPS > 0 인 경우만 계산 (EPS 음수/0이면 의미 없음)
    df["배당성향(%)"] = np.where(
        (df["EPS"] > 0) & (df["DPS_최근"] > 0),
        (df["DPS_최근"] / df["EPS"]) * 100,
        np.nan,
    )
    # 배당_경고신호 — Value Trap(주가폭락 착시) 또는 Payout Trap(이익훼손 배당) 감지
    _rs_weak = (df["RS_등급"].fillna(50) < 30) if "RS_등급" in df.columns else False
    df["배당_경고신호"] = np.where(
        (df["배당성향(%)"].fillna(0) > 80)
        | ((df["배당수익률(%)"] > 10) & _rs_weak)
        | (df["현금전환율(%)"].fillna(100) < 70),
        1,
        0,
    ).astype(int)
    # S-RIM 동적 할인율: config 설정값 사용 (환경변수로 재정의 가능)
    # Ke = 무위험수익률(국고채 3Y) + 시장위험프리미엄
    Ke = config.RISK_FREE_RATE + config.EQUITY_RISK_PREMIUM  # default: 3.5 + 5.5 = 9.0%
    # 지속계수 ω=0.9: 초과이익의 점진적 소멸 반영 (영구 지속 가정 방지)
    # 공식: BPS + BPS * (ROE - Ke) * ω / (1 + Ke/100 - ω)
    omega = 0.9
    Ke_dec = Ke / 100.0
    srim_denom = max(1 + Ke_dec - omega, 1e-6)  # 0 나누기 방지
    df["적정주가_SRIM"] = np.where(
        (df["ROE(%)"] > Ke) & (df["BPS"] > 0),
        df["BPS"] + df["BPS"] * (df["ROE(%)"] / 100.0 - Ke_dec) * omega / srim_denom,
        df["BPS"],  # ROE ≤ Ke: 초과이익 없음 → BPS 그대로 (할인 없음, 과대평가 방지)
    )
    df["괴리율(%)"] = ((df["적정주가_SRIM"] - df["종가"]) / df["종가"]) * 100

    # PER 이상치 플래그
    df["PER_이상"] = np.where(
        pd.notna(df["PER"]) & ((df["PER"] < 0.5) | (df["PER"] > 500)), 1, 0
    ).astype(int)

    # 이자보상배율: analyze_one_stock에서 산출된 값 그대로 사용 (열 없으면 NaN으로 초기화)
    if "이자보상배율" not in df.columns:
        df["이자보상배율"] = np.nan

    # F7 업데이트: 발행주식수 미증가 (shares_df 기반)
    if shares_df is not None and not shares_df.empty:
        shares_by_code = shares_df.sort_values("기준일").groupby("종목코드")["발행주식수"]
        for idx, row in df.iterrows():
            code = row["종목코드"]
            if code in shares_by_code.groups:
                s_vals = shares_by_code.get_group(code).dropna().values
                if len(s_vals) >= 2:
                    f7_val = 1 if s_vals[-1] <= s_vals[-2] else 0
                    df.at[idx, "F7_희석없음"] = f7_val
                    df.at[idx, "F스코어"] = (
                        row.get("F1_수익성", 0) + row.get("F2_영업CF", 0)
                        + row.get("F3_ROA개선", 0) + row.get("F4_이익품질", 0)
                        + row.get("F5_레버리지", 0) + row.get("F6_유동성", 0)
                        + f7_val
                        + row.get("F8_매출총이익률", 0) + row.get("F9_자산회전율", 0)
                    )

    return df

def calc_technical_indicators(df, price_hist, index_hist=None, master=None):
    """기술적 지표 계산. index_hist/master 제공 시 Composite RS도 계산.

    Args:
        df: 종목 데이터프레임 (종목코드 포함)
        price_hist: 주가 히스토리 (종목코드, 날짜, 종가, 거래량 등)
        index_hist: 지수 히스토리 (지수코드, 날짜, 종가) — KOSPI/KOSDAQ
        master: 종목 마스터 (종목코드, 시장구분) — RS 계산 시 지수 선택에 사용
    """
    if price_hist.empty: return df

    # 지수 데이터 사전 준비 (날짜 → 종가 매핑)
    idx_map = {}  # {"KOSPI": pd.Series(종가, index=날짜), "KOSDAQ": ...}
    if index_hist is not None and not index_hist.empty:
        for idx_code in ["KOSPI", "KOSDAQ"]:
            sub = index_hist[index_hist["지수코드"] == idx_code].copy()
            sub["날짜"] = pd.to_datetime(sub["날짜"])
            sub = sub.sort_values("날짜").set_index("날짜")["종가"]
            idx_map[idx_code] = sub

    # FDR Fallback: DB에 지수 데이터 없으면 FinanceDataReader로 즉시 다운로드
    if not idx_map:
        try:
            import FinanceDataReader as fdr
            from datetime import datetime, timedelta
            _start = (datetime.today() - timedelta(days=730)).strftime("%Y-%m-%d")
            for idx_code, fdr_code in [("KOSPI", "KS11"), ("KOSDAQ", "KQ11")]:
                try:
                    _data = fdr.DataReader(fdr_code, start=_start)
                    if _data is not None and not _data.empty:
                        _data.index = pd.to_datetime(_data.index)
                        idx_map[idx_code] = _data["Close"].sort_index()
                except Exception:
                    pass
            if idx_map:
                log.info("지수 데이터 FDR fallback 적용 (KOSPI/KOSDAQ)")
        except ImportError:
            pass

    # 종목 → 시장구분 매핑
    market_map = {}
    if master is not None and not master.empty and "시장구분" in master.columns:
        market_map = master.drop_duplicates("종목코드").set_index("종목코드")["시장구분"].to_dict()

    def _rs_ret(price_series, n):
        """날짜 인덱스 기반 n거래일 전 대비 수익률(%). shift 금지 — iloc 기반."""
        if len(price_series) < n + 1:
            return np.nan
        p_now = price_series.iloc[-1]
        p_prev = price_series.iloc[-(n + 1)]
        if p_prev <= 0:
            return np.nan
        return (p_now / p_prev - 1) * 100

    def _index_ret(idx_series, stock_dates, n):
        """종목 날짜 기준으로 지수 수익률 계산 (날짜 merge 방식)."""
        if idx_series is None or len(idx_series) == 0:
            return np.nan
        stock_dates_dt = pd.to_datetime(stock_dates)
        # 종목의 최신/n거래일전 날짜와 가장 가까운 지수 날짜 사용
        latest_date = stock_dates_dt.max()
        idx_sorted = idx_series.sort_index()
        # 최신 날짜 지수값
        available = idx_sorted.index[idx_sorted.index <= latest_date]
        if len(available) == 0:
            return np.nan
        p_now = idx_sorted[available[-1]]
        # n거래일 전: 종목 기준 n번째 이전 날짜
        if len(stock_dates_dt) < n + 1:
            return np.nan
        target_date = sorted(stock_dates_dt)[-(n + 1)]
        available_prev = idx_sorted.index[idx_sorted.index <= target_date]
        if len(available_prev) == 0:
            return np.nan
        p_prev = idx_sorted[available_prev[-1]]
        if p_prev <= 0:
            return np.nan
        return (p_now / p_prev - 1) * 100

    ph_grouped = {k: v for k, v in price_hist.groupby("종목코드")} if not price_hist.empty else {}
    techs = []
    for code in df["종목코드"].unique():
        if code not in ph_grouped: continue
        ph = ph_grouped[code].sort_values("날짜")
        if len(ph) < 5: continue
        close = ph["종가"].iloc[-1]
        ma20 = ph["종가"].rolling(20).mean().iloc[-1] if len(ph)>=20 else np.nan
        ma60 = ph["종가"].rolling(60).mean().iloc[-1] if len(ph)>=60 else np.nan
        h52, l52 = ph["종가"].max(), ph["종가"].min()

        rsi = np.nan
        if len(ph) >= 15:
            delta = ph["종가"].diff()
            gain = (delta.where(delta > 0, 0)).rolling(14).mean().iloc[-1]
            loss = (-delta.where(delta < 0, 0)).rolling(14).mean().iloc[-1]
            rsi = 100 - (100 / (1 + (gain / loss))) if loss > 0 else 100

        # 거래대금 컬럼 존재하고 실제 값이 있으면 사용, 아니면 종가×거래량으로 계산
        v_col = "거래대금" if ("거래대금" in ph.columns and ph["거래대금"].notna().any()) else None
        v20 = (ph[v_col].tail(20).mean() if v_col else (ph["종가"]*ph["거래량"]).tail(20).mean()) if len(ph)>=20 else np.nan
        v5 = (ph[v_col].tail(5).mean() if v_col else (ph["종가"]*ph["거래량"]).tail(5).mean())
        vol60 = ph["종가"].pct_change().tail(60).std() * np.sqrt(252) * 100 if len(ph)>=60 else np.nan

        # ── Raw RS (기간별 초과수익률) — composite는 루프 후 선랭킹 방식으로 계산 ──
        rs_60d = rs_120d = rs_250d = np.nan
        if idx_map:
            market = market_map.get(str(code), "KOSPI")
            if market not in idx_map:
                market = "KOSPI"
            idx_s = idx_map[market]
            stock_dates = ph["날짜"]

            stock_ret_60 = _rs_ret(ph["종가"], 60)
            stock_ret_120 = _rs_ret(ph["종가"], 120)
            stock_ret_250 = _rs_ret(ph["종가"], 250)

            idx_ret_60 = _index_ret(idx_s, stock_dates, 60)
            idx_ret_120 = _index_ret(idx_s, stock_dates, 120)
            idx_ret_250 = _index_ret(idx_s, stock_dates, 250)

            if pd.notna(stock_ret_60) and pd.notna(idx_ret_60):
                rs_60d = stock_ret_60 - idx_ret_60
            if pd.notna(stock_ret_120) and pd.notna(idx_ret_120):
                rs_120d = stock_ret_120 - idx_ret_120
            if pd.notna(stock_ret_250) and pd.notna(idx_ret_250):
                rs_250d = stock_ret_250 - idx_ret_250

        techs.append({
            "종목코드": code, "RSI_14": rsi, "MA20_이격도(%)": (close/ma20-1)*100 if pd.notna(ma20) else np.nan,
            "MA60_이격도(%)": (close/ma60-1)*100 if pd.notna(ma60) else np.nan,
            "52주_최고대비(%)": (close/h52-1)*100, "52주_최저대비(%)": (close/l52-1)*100,
            "거래대금_20일평균": v20, "거래대금_증감(%)": (v5/v20-1)*100 if pd.notna(v20) and v20 > 0 else np.nan,
            "변동성_60일(%)": vol60,
            "RS_60d": rs_60d, "RS_120d": rs_120d, "RS_250d": rs_250d,
        })

    tech_df = pd.DataFrame(techs)

    # ── Composite RS: 기간별 선랭킹 후 가중합산 (O'Neil 방식) ──
    # 각 기간의 raw RS를 전 종목 대비 percentile rank(0~100)로 먼저 변환 →
    # 이후 가중평균: 수익률 크기 편향 제거, 기간별 균등 기여 보장
    _rs_weights = [("RS_60d", 0.4), ("RS_120d", 0.3), ("RS_250d", 0.3)]
    for rs_col, _ in _rs_weights:
        if rs_col in tech_df.columns and tech_df[rs_col].notna().any():
            tech_df[f"_rank_{rs_col}"] = tech_df[rs_col].rank(pct=True, na_option="keep") * 100
        else:
            tech_df[f"_rank_{rs_col}"] = np.nan

    def _weighted_composite(row):
        total_val, total_w = 0.0, 0.0
        for rs_col, w in _rs_weights:
            v = row.get(f"_rank_{rs_col}", np.nan)
            if pd.notna(v):
                total_val += v * w
                total_w += w
        return total_val / total_w if total_w > 0 else np.nan

    tech_df["Composite_RS"] = tech_df.apply(_weighted_composite, axis=1)
    # 임시 랭킹 컬럼 제거
    tech_df = tech_df.drop(columns=[f"_rank_{c}" for c, _ in _rs_weights if f"_rank_{c}" in tech_df.columns])

    result = df.merge(tech_df, on="종목코드", how="left")

    # RS_등급: Composite_RS의 전체 종목 백분위 (0~100)
    if "Composite_RS" in result.columns and result["Composite_RS"].notna().any():
        rs_rank = result["Composite_RS"].rank(pct=True, na_option="keep") * 100
        result["RS_등급"] = rs_rank
    else:
        result["RS_등급"] = np.nan

    return result

def calc_investor_strength(inv_df, daily, price_hist=None):
    """수급강도 + 스마트머니 매집 연속성 + VCP 신호 계산.

    Args:
        inv_df: 투자자 매매동향 (종목코드, 날짜, 외국인순매수, 기관순매수)
        daily: 일별 시세 (종목코드, 시가총액)
        price_hist: 주가 히스토리 (종목코드, 날짜, 종가, 거래량) — VCP 계산에 필요
    """
    if inv_df.empty: return pd.DataFrame(columns=["종목코드", "수급강도", "외인순매수_20d", "기관순매수_20d"])
    res = []
    
    inv_grouped = {k: v for k, v in inv_df.groupby("종목코드")}
    daily_mcap = daily.drop_duplicates("종목코드").set_index("종목코드")["시가총액"].to_dict()
    ph_grouped = {k: v for k, v in price_hist.groupby("종목코드")} if price_hist is not None and not price_hist.empty else {}

    for code in inv_df["종목코드"].unique():
        if code not in inv_grouped: continue
        df_code = inv_grouped[code].sort_values("날짜", ascending=False).head(20)
        f_sum, i_sum = df_code["외국인순매수"].sum(), df_code["기관순매수"].sum()
        mcap = daily_mcap.get(code, np.nan)
        strength = ((f_sum + i_sum) / mcap) * 100 if pd.notna(mcap) and mcap > 0 else np.nan

        # ── 스마트머니 매집 연속성 ──
        n = len(df_code)
        if n > 0:
            buy_days = ((df_code["외국인순매수"] > 0) | (df_code["기관순매수"] > 0)).sum()
            both_days = ((df_code["외국인순매수"] > 0) & (df_code["기관순매수"] > 0)).sum()
            smart_ratio = buy_days / n
            both_ratio = both_days / n
        else:
            smart_ratio = np.nan
            both_ratio = np.nan

        # ── VCP 신호 (가격 + 거래량 축소 동시 확인) ──
        vcp = 0
        if ph_grouped and code in ph_grouped:
            ph = ph_grouped[code].sort_values("날짜")
            if len(ph) >= 60:
                close_s = ph["종가"]
                # 가격 CV(변동계수) 비교
                cv20 = close_s.tail(20).std() / close_s.tail(20).mean() if close_s.tail(20).mean() > 0 else np.nan
                cv60 = close_s.tail(60).std() / close_s.tail(60).mean() if close_s.tail(60).mean() > 0 else np.nan
                price_compress = pd.notna(cv20) and pd.notna(cv60) and cv20 < cv60
                # 거래량 축소 비교
                vol_col = "거래량"
                if vol_col in ph.columns:
                    vol20 = ph[vol_col].tail(20).mean()
                    vol60 = ph[vol_col].tail(60).mean()
                    vol_compress = pd.notna(vol20) and pd.notna(vol60) and vol60 > 0 and vol20 < vol60
                else:
                    vol_compress = False
                # 스마트머니 승률 60%+ AND 가격+거래량 동시 축소
                sm_ok = pd.notna(smart_ratio) and smart_ratio >= 0.6
                vcp = 1 if (price_compress and vol_compress and sm_ok) else 0

        res.append({
            "종목코드": code,
            "수급강도": strength,
            "외인순매수_20d": f_sum,
            "기관순매수_20d": i_sum,
            "스마트머니_승률": smart_ratio,
            "양매수_비율": both_ratio,
            "VCP_신호": vcp,
        })
    return pd.DataFrame(res)

# ═════════════════════════════════════════════
# 스코어링 & 저장 (v8 스타일 유지)
# ═════════════════════════════════════════════

def calc_strategy_scores(df):
    def get_rank(col, asc=True, zero_if_nan=False):
        """백분위 랭킹 (0~100). zero_if_nan=True 시 NaN 종목은 0점 처리."""
        if col not in df.columns:
            return pd.Series(0.0 if zero_if_nan else 50.0, index=df.index)
        series = df[col].copy()
        if zero_if_nan:
            nan_mask = series.isna()
            # NaN을 최하위로 채운 뒤 랭킹, 이후 0으로 덮어쓰기
            fill_val = series.min() - 1 if (asc and series.notna().any()) else series.max() + 1 if series.notna().any() else 0
            series = series.fillna(fill_val)
            ranked = series.rank(pct=True) * 100 if asc else (1 - series.rank(pct=True)) * 100
            ranked[nan_mask] = 0.0
            return ranked
        else:
            series = series.fillna(series.median() if not series.isna().all() else 0)
            return series.rank(pct=True) * 100 if asc else (1 - series.rank(pct=True)) * 100

    # ── 턴어라운드 프리미엄: 흑자전환 종목의 NaN CAGR을 75th percentile로 대체 ──
    # 적자→흑자 전환 시 CAGR이 NaN(음수 기저)이 되어 0점 처리되는 문제 완화
    # 75th percentile = 보수적 상위 25% 인정 (과대평가 방지)
    if "순이익_CAGR" in df.columns and "흑자전환" in df.columns:
        _cagr_p75 = df["순이익_CAGR"].quantile(0.75) if df["순이익_CAGR"].notna().any() else 0
        _turnaround_mask = (df["흑자전환"] == 1) & df["순이익_CAGR"].isna()
        df["_순이익_CAGR_adj"] = df["순이익_CAGR"].copy()
        df.loc[_turnaround_mask, "_순이익_CAGR_adj"] = _cagr_p75
    else:
        df["_순이익_CAGR_adj"] = df.get("순이익_CAGR", pd.Series(np.nan, index=df.index))

    # ── 성장률 Winsorization: 극단값(소형주 base effect 등)이 백분위 왜곡 방지 ──
    # PEG 계산은 별도로 100% 캡 적용(line ~688). 여기선 스코어링용 캡 처리.
    _CAGR_CAP = 150  # % 상한 — 150% 초과 성장은 일회성/소형주 base effect로 간주
    for _gcol in ["매출_CAGR", "영업이익_CAGR", "순이익_CAGR"]:
        if _gcol in df.columns:
            df[_gcol] = df[_gcol].clip(upper=_CAGR_CAP)
    _YOY_CAP = 300  # 분기 YoY는 베이스가 더 작아 상한을 넓게
    for _gcol in ["Q_매출_YoY(%)", "Q_영업이익_YoY(%)", "Q_순이익_YoY(%)"]:
        if _gcol in df.columns:
            df[_gcol] = df[_gcol].clip(upper=_YOY_CAP)

    # NaN=0점: PER/PBR/PEG (낮을수록 좋은 지표, NaN=적자/음자본 → 최하위)
    # NaN=0점: ROE, FCF수익률 (데이터 없으면 점수 불허)
    S_PER_inv = get_rank("PER", False, zero_if_nan=True)
    S_PBR_inv = get_rank("PBR", False, zero_if_nan=True)
    S_PEG_inv = get_rank("PEG", False, zero_if_nan=True)
    S_ROE = get_rank("ROE(%)", asc=True, zero_if_nan=True)
    S_FCF = get_rank("FCF수익률(%)", asc=True, zero_if_nan=True)
    S_OpCAGR, S_QOpYoY = get_rank("영업이익_CAGR"), get_rank("Q_영업이익_YoY(%)")
    S_Div, S_Supply, S_Vol = get_rank("배당수익률(%)"), get_rank("수급강도"), get_rank("거래대금_20일평균")

    # 핵심 모멘텀/퀄리티 지표 — NaN은 0점 처리 (증명 안 된 종목에 점수 불허)
    S_ROIC = get_rank("ROIC(%)", asc=True, zero_if_nan=True)
    S_RS = get_rank("RS_등급", asc=True, zero_if_nan=True)
    S_Accel = get_rank("실적가속_연속", asc=True, zero_if_nan=True)
    S_SmartMoney = get_rank("스마트머니_승률", asc=True, zero_if_nan=True)
    S_GPM_delta = get_rank("GPM_변화(pp)", asc=True, zero_if_nan=True)
    # 주도주: RS_등급(25%) + 수급강도(20%) + 거래대금(10%) + 영업이익CAGR(15%) + Q_YoY(15%) + 실적가속(10%) + RSI(5%)
    df["주도주_점수"] = (S_RS*0.25 + S_Supply*0.20 + S_Vol*0.10 + S_OpCAGR*0.15 + S_QOpYoY*0.15 + S_Accel*0.10 + get_rank("RSI_14")*0.05)
    # FCF수익률(25%=Value) + ROIC(25%=Quality) + F스코어(20%=Health) + 괴리율(20%=MoS) + PEG역순(10%=Growth-Value)
    df["우량가치_점수"] = (S_FCF*0.25 + S_ROIC*0.25 + get_rank("F스코어")*0.20 + get_rank("괴리율(%)")*0.20 + S_PEG_inv*0.10)
    # Q_영업이익_YoY(20%) + 실적가속_연속(20%) + 영업이익_CAGR(15%) + RS_등급(25%) + PEG역순(20%)
    df["고성장_점수"] = (S_QOpYoY*0.20 + S_Accel*0.20 + S_OpCAGR*0.15 + S_RS*0.25 + S_PEG_inv*0.20)
    # 현금배당: FCF(25%) + 배당수익률(20%) + DPS성장(15%) + ROIC해자(15%) + 배당성향역순(10%) + F스코어(10%) + 부채비율(5%)
    # + 배당연속증가 보너스(로그 스케일, 최대+10) + 수익동반증가 추가보너스(+2) × 경고신호 페널티 승수(×0.7)
    # 기존: 배당_수익동반증가 binary * 5 → 1년=10년 동일 +5점 문제
    # 개선: log2(연수+1)*3 → 1년=+2.1, 3년=+4.5, 5년=+5.6, 10년=+7.5, 상한+10
    S_PayoutInv = get_rank("배당성향(%)", asc=False, zero_if_nan=False)  # 낮을수록 good
    _raw_div = (
        S_FCF * 0.25
        + S_Div * 0.20
        + get_rank("DPS_CAGR") * 0.15
        + S_ROIC * 0.15
        + S_PayoutInv * 0.10
        + get_rank("F스코어") * 0.10
        + get_rank("부채비율(%)", False) * 0.05
    )
    _consec = df.get("배당_연속증가", pd.Series(0, index=df.index)).fillna(0).clip(lower=0)
    _div_bonus = np.minimum(np.log2(_consec.where(_consec > 0, np.nan) + 1) * 3, 10).fillna(0)
    _div_bonus += df.get("배당_수익동반증가", pd.Series(0, index=df.index)).fillna(0) * 2
    _div_penalty = np.where(df["배당_경고신호"] == 1, 0.7, 1.0)
    df["현금배당_점수"] = (_raw_div + _div_bonus) * _div_penalty
    # 턴어라운드 (Grand Master 개편): 이익률변동폭(10%) + 흑자전환(15%) + 스마트머니(15%)
    #   + GPM변화(10%) + Q_매출_YoY(15%) + 이자보상배율(10%) + 퀄리티_턴어라운드(15%) + 괴리율(10%)
    # [기각] F스코어 제거 — 퀄리티_턴어라운드(GPM+OCF+ROIC)가 핵심 구성요소 대체
    # [기각] Q_매출_YoY >0 하드필터 — 구조조정형 매출 일시 감소 허용 위해 점수만 반영
    S_Sales_YoY    = get_rank("Q_매출_YoY(%)",     asc=True, zero_if_nan=True)
    S_Interest_Cov = get_rank("이자보상배율",       asc=True, zero_if_nan=True)
    S_Qual_Turn    = get_rank("퀄리티_턴어라운드",  asc=True, zero_if_nan=True)
    df["턴어라운드_점수"] = (
        get_rank("이익률_변동폭") * 0.10   # 이익률 개선 폭
        + get_rank("흑자전환")   * 0.15    # 흑자전환 상징성
        + S_SmartMoney           * 0.15    # 수급 선취매
        + S_GPM_delta            * 0.10    # 원가 경쟁력
        + S_Sales_YoY            * 0.15    # [신규] 탑라인(매출) 회복 검증
        + S_Interest_Cov         * 0.10    # [신규] 파산 리스크 통제
        + S_Qual_Turn            * 0.15    # [신규] GPM+OCF+ROIC 복합 품질 신호
        + get_rank("괴리율(%)")  * 0.10    # S-RIM 안전마진
    )  # 합계: 10+15+15+10+15+10+15+10 = 100%

    # ── 종합점수: 성장성 / 안정성 / 주가 세 축 균형 (각 0-100, 합산 평균 → 0-100) ──
    # 성장성 (33%): 영업이익CAGR 35% + 매출CAGR 30% + 최근분기YoY 25% + 실적가속도 10%
    성장성_점수 = S_OpCAGR * 0.35 + get_rank("매출_CAGR") * 0.30 + S_QOpYoY * 0.25 + S_Accel * 0.10
    # 안정성 (33%): ROE 40% + F스코어(재무건전성) 35% + FCF수익률 25%
    안정성_점수 = S_ROE * 0.40 + get_rank("F스코어") * 0.35 + S_FCF * 0.25
    # 주가 (33%): PER역순 40% + S-RIM괴리율 35% + PBR역순 25%
    가격_점수   = S_PER_inv * 0.40 + get_rank("괴리율(%)") * 0.35 + S_PBR_inv * 0.25

    df["성장성_점수"] = 성장성_점수
    df["안정성_점수"] = 안정성_점수
    df["가격_점수"]   = 가격_점수
    df["종합점수"]    = (성장성_점수 + 안정성_점수 + 가격_점수) / 3

    # 임시 컬럼 정리
    if "_순이익_CAGR_adj" in df.columns:
        df = df.drop(columns=["_순이익_CAGR_adj"])
    return df

def apply_leaders_screen(df):
    # 시장 주도주: 대형주 + 유동성 + 수익성 + RS 상위 20% + 수급강도 양수
    mask = (df["시가총액"]>=100_000_000_000) & (df["TTM_순이익"]>0) & (df["주도주_점수"]>0)
    # RS_등급 상위 20% (기존 70 → 80)
    if "RS_등급" in df.columns and df["RS_등급"].notna().any():
        mask = mask & ((df["RS_등급"].fillna(0) >= 80) | (df["RS_등급"].isna()))
    # 거래대금 5억 이상 (기존 1억 → 5억)
    if "거래대금_20일평균" in df.columns:
        mask = mask & ((df["거래대금_20일평균"] > 500_000_000) | df["거래대금_20일평균"].isna())
    # 수급강도 양수 (외국인+기관 순매수)
    if "수급강도" in df.columns:
        mask = mask & (df["수급강도"].fillna(0) > 0)
    return df[mask].sort_values("주도주_점수", ascending=False)

def apply_quality_value_screen(df):
    # 금융주/지주사 판별: 종목명 키워드 OR 유동비율 데이터 없음(금융업 구조적 특성)
    is_finance = (
        df["종목명"].str.contains("지주|금융|은행|증권|생명|화재", na=False)
        | df["유동비율(%)"].isna()
        | (df["유동비율(%)"].fillna(0) == 0)
    )
    # Track A: 일반 기업 (제조/서비스) — 육각형 우량 룰
    mask_general = (
        (~is_finance)
        & (df["ROIC(%)"].fillna(0) >= 10)                   # 그린블라트: ROIC 10%+ 허들 (12→10 완화)
        & (df["F스코어"].fillna(0) >= 5)                    # 피오트로스키: 9점 중 5점 이상 (F3/F4가 현금흐름 검증 포함)
        & (df["PEG"].fillna(99) < 1.2)                     # 피터 린치: 성장 대비 저평가 (NaN은 검증불가로 제외)
        & (df["부채비율(%)"].fillna(999) < 120)             # 레버리지 제한
        & (df["유동비율(%)"].fillna(0) > 120)               # 단기 파산 리스크 차단
        & (df["순이익_당기양수"].fillna(0) == 1)            # 당기 흑자
        & (df["순이익_전년음수"].fillna(0) == 0)            # 전년도 흑자 (연속 흑자)
        & (df["시가총액"].fillna(0) >= 100_000_000_000)     # 1000억 이상
    )
    # Track B: 금융주/지주사 전용 (버핏의 은행주 선별 잣대)
    mask_finance = (
        is_finance
        & (df["ROE(%)"].fillna(0) >= 8)                    # 금융주: ROIC 대신 ROE 방어력
        & (df["PBR"].fillna(99) < 1.5)                     # 한국 금융주 현실 감안 (1.0→1.5 완화)
        & (df["배당수익률(%)"].fillna(0) >= 2.0)            # 주주환원: 현금배당 최소 2% (한국 금융주 현실)
        & (df["F스코어"].fillna(0) >= 4)                   # 유동성 왜곡 감안 4점으로 완화
        & (df["순이익_당기양수"].fillna(0) == 1)            # 연속 흑자 필수
        & (df["순이익_전년음수"].fillna(0) == 0)
        & (df["시가총액"].fillna(0) >= 300_000_000_000)     # 3000억 이상 (금융 소형주 제거)
    )
    return df[mask_general | mask_finance].sort_values("우량가치_점수", ascending=False)

def apply_growth_mom_screen(df):
    mask = (
        # 1. 외형과 내실의 동반 성장 (AND 유지, 기준 10%로 현실화)
        (df["매출_CAGR"].fillna(0) >= 10)
        & (df["영업이익_CAGR"].fillna(0) >= 10)
        # 2. 최근 분기 실적 성장 (역성장 제외)
        & (df["Q_영업이익_YoY(%)"].fillna(0) > 0)
        # 3. 시장 주도 모멘텀 (70 -> 50: 시장 평균 이상이면 후보군 포함)
        & (df["RS_등급"].fillna(0) >= 50)
        # 4. 리스크 방어 (흑자도산 방지)
        & (df["TTM_영업CF"].fillna(-1) > 0)
        & (df["시가총액"].fillna(0) >= 50_000_000_000)
    )
    return df[mask].sort_values("고성장_점수", ascending=False)

def apply_cash_div_screen(df):
    mask = (
        (df.get("FCF수익률(%)", 0) >= 3)
        & (df["배당수익률(%)"] >= 1)
        & (df["부채비율(%)"].fillna(999) < 120)       # 150 → 120 강화 (재무 안전성)
        & (df["시가총액"] >= 50_000_000_000)          # 500억 유지
        & (df["배당성향(%)"].fillna(999) < 80)         # 신규: EPS 대비 과도 배당 차단
        & (df["현금전환율(%)"].fillna(0) >= 70)        # 신규: 이익→현금 전환 품질
    )
    return df[mask].sort_values("현금배당_점수", ascending=False)

def apply_turnaround_screen(df):
    base_mask = (
        ((df.get("흑자전환")==1) | (df.get("이익률_급개선")==1))
        & (df["TTM_순이익"] > 0)
        & (df["TTM_영업CF"].fillna(-1) > 0)         # [신규] 회계 흑자가 아닌 실제 현금 창출 검증
        & (df["Q_매출_YoY(%)"].fillna(0) > -15)     # [신규] 매출 급감(-15%↓) 방지 (구조조정형 허용)
        & (df["시가총액"] >= 30_000_000_000)
        & (df["이자보상배율"].fillna(0) > 1.5)       # [신규] 이자 상환 능력 (파산 리스크 통제)
    )
    # 스마트머니 승률 50%+ OR VCP 신호 보조 조건 (있을 때만 적용)
    if "스마트머니_승률" in df.columns:
        smart_mask = (df["스마트머니_승률"].fillna(0) >= 0.5) | (df.get("VCP_신호", 0).fillna(0) == 1)
        mask = base_mask & smart_mask
        # 스마트머니 데이터 없는 종목은 기본 조건으로 통과
        no_data_mask = base_mask & df["스마트머니_승률"].isna()
        mask = mask | no_data_mask
    else:
        mask = base_mask
    return df[mask].sort_values("턴어라운드_점수", ascending=False)

def save_to_excel(df, filepath, sheet_name="Result"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.numbers import FORMAT_NUMBER_00
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = sheet_name
    col_groups = {
        "기본정보": ["종목코드", "종목명", "종가", "시가총액", "상장주식수", "데이터_연수"],
        "주요지표": ["PER", "PBR", "PSR", "PEG", "PER_이상", "ROE(%)", "EPS", "BPS",
                   "부채비율(%)", "유동비율(%)", "영업이익률(%)", "이익수익률(%)", "FCF수익률(%)",
                   "배당수익률(%)", "이익품질_양호", "이자보상배율", "현금전환율(%)", "CAPEX비율(%)"],
        "F스코어": ["F스코어", "F1_수익성", "F2_영업CF", "F3_ROA개선", "F4_이익품질", "F5_레버리지", "F6_유동성", "F7_희석없음", "F8_매출총이익률", "F9_자산회전율"],
        "수급/거래": ["수급강도", "외인순매수_20d", "기관순매수_20d", "거래대금_20일평균", "거래대금_증감(%)"],
        "점수": ["종합점수", "성장성_점수", "안정성_점수", "가격_점수", "주도주_점수", "우량가치_점수", "고성장_점수", "현금배당_점수", "턴어라운드_점수"],
        "성장추세": ["매출_CAGR", "영업이익_CAGR", "순이익_CAGR", "매출_연속성장", "영업이익_연속성장", "이익률_변동폭", "흑자전환", "순이익_전년음수", "순이익_당기양수"],
        "밸류에이션": ["적정주가_SRIM", "괴리율(%)"],
        "Forward추정치": ["컨센서스_커버리지", "Fwd_PER", "Fwd_PBR", "Fwd_EPS", "Fwd_ROE(%)", "Fwd_OPM(%)",
                         "Fwd_영업이익_성장률(%)", "Fwd_매출_성장률(%)", "Fwd_순이익_성장률(%)", "Fwd_2yr_영업이익_성장(%)"],
    }
    # NaN을 None(빈칸)으로 유지하되 숫자형 셀에 별도 포맷 적용
    # — 수식이 성립하지 않는 지표(현금전환율, CAPEX비율 등)는 셀을 비워두고
    #   조건부 서식으로 빈 셀에 "-" 표시 (데이터 타입 보존)
    _nan_fmt_cols = {"현금전환율(%)", "CAPEX비율(%)", "부채상환능력", "이자보상배율",
                     "PEG", "PER", "PBR", "PSR", "FCF수익률(%)", "이익수익률(%)"}
    ordered_cols = []
    for g in col_groups.values():
        for c in g:
            if c in df.columns: ordered_cols.append(c)
    for c in df.columns:
        if c not in ordered_cols and not c.startswith("S_"): ordered_cols.append(c)
    export_df = df[ordered_cols].copy()
    fills = {
        "기본정보": PatternFill("solid", fgColor="D6E4F0"), "주요지표": PatternFill("solid", fgColor="E2EFDA"),
        "F스코어": PatternFill("solid", fgColor="FCE4D6"), "수급/거래": PatternFill("solid", fgColor="FDE9D9"),
        "점수": PatternFill("solid", fgColor="C6EFCE"), "성장추세": PatternFill("solid", fgColor="FFF2CC"),
        "밸류에이션": PatternFill("solid", fgColor="DAEEF3"), "Forward추정치": PatternFill("solid", fgColor="E8D5F5")
    }
    for col_idx, col_name in enumerate(ordered_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        for grp, cols in col_groups.items():
            if col_name in cols: cell.fill = fills[grp]; break
    # 헤더에서 NaN 포맷 대상 컬럼의 인덱스 미리 파악
    _nan_col_indices = {
        col_idx for col_idx, col_name in enumerate(ordered_cols, 1)
        if col_name in _nan_fmt_cols
    }
    for row_idx, (_, row_data) in enumerate(export_df.iterrows(), 2):
        for col_idx, col_name in enumerate(ordered_cols, 1):
            val = row_data[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(val):
                # NaN 포맷 대상 컬럼: "-" 문자열로 표시 (정렬은 우측)
                if col_idx in _nan_col_indices:
                    cell.value = "-"
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.value = None
            else:
                if isinstance(val, (float, np.floating)):
                    cell.value = round(float(val), 2)
                else:
                    cell.value = val
    ws.freeze_panes = "C2"
    wb.save(filepath); log.info(f"Saved: {filepath}")

def run():
    daily, fs, ind = load_table("daily"), load_table("financial_statements"), load_table("indicators")
    shares, hist, inv = load_table("shares"), load_table("price_history"), load_table("investor_trading")
    index_hist = load_table("index_history")
    master = load_table("master")
    if daily.empty: return
    ind = preprocess_indicators(ind); mult = detect_unit_multiplier(ind); anal = analyze_all(fs, ind)
    full = calc_valuation(daily, anal, mult, shares)
    full = calc_technical_indicators(
        full, hist,
        index_hist=index_hist if not index_hist.empty else None,
        master=master if not master.empty else None,
    )
    full = full.merge(
        calc_investor_strength(inv, daily, price_hist=hist if not hist.empty else None),
        on="종목코드", how="left",
    )
    full = calc_strategy_scores(full)
    save_to_excel(full.sort_values("종합점수", ascending=False), DATA_DIR / "quant_all_stocks.xlsx", "All")
    save_to_excel(apply_leaders_screen(full), DATA_DIR / "quant_leaders.xlsx", "Leaders")
    save_to_excel(apply_quality_value_screen(full), DATA_DIR / "quant_quality_value.xlsx", "QualityValue")
    save_to_excel(apply_growth_mom_screen(full), DATA_DIR / "quant_growth_mom.xlsx", "GrowthMom")
    save_to_excel(apply_cash_div_screen(full), DATA_DIR / "quant_cash_div.xlsx", "CashDiv")
    save_to_excel(apply_turnaround_screen(full), DATA_DIR / "quant_turnaround.xlsx", "Turnaround")

if __name__ == "__main__": run()
